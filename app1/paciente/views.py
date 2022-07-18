from django.forms import modelform_factory
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl import Workbook
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect

# Create your views here.
from openpyxl.worksheet.table import Table, TableStyleInfo

from paciente.forms import PacienteFormulario
from paciente.models import Paciente

def detalle_paciente(request, id):
    # paciente = Paciente.objects.get(pk=id)
    paciente = get_object_or_404(Paciente, pk=id)
    mensaje_pac = {'paciente': paciente}
    return render(request, 'detallepaciente.html', mensaje_pac)

#PacienteFormulario = modelform_factory(Paciente, exclude=[])
def nuevo_paciente(request):
    if request.POST:
        formPaciente = PacienteFormulario(request.POST)
        if formPaciente.is_valid():
            formPaciente.save()
            return redirect('inicio')
    else:
        formPaciente = PacienteFormulario()
    return render(request, 'nuevopaciente.html', {'formPaciente': formPaciente})

def modificar_paciente(request, id):
    paciente = get_object_or_404(Paciente, pk=id)
    if request.method == 'POST':
        formPaciente = PacienteFormulario(request.POST, instance=paciente)
        if formPaciente.is_valid():
            formPaciente.save()
            return redirect('inicio')
    else:
        formPaciente = PacienteFormulario(instance=paciente)
    mensaje = {'formPaciente': formPaciente}
    return render(request, 'modificarpaciente.html', context=mensaje)

def eliminar_paciente(request, id):
    paciente = get_object_or_404(Paciente, pk=id)
    if paciente:
        paciente.delete()
    return redirect('inicio')

def reporte_paciente(request):
    pacientes = Paciente.objects.order_by('apellidos')
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('B1:F1')
    cell = ws.cell(row=1, column=2)
    cell.value = 'REPORTE DE PACIENTES DE LA CLINICA MEDICLINIC'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = cell.font.copy(bold=True)
    ws.title = 'Reporte de Pacientes'
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 35
    ws['B3'] = 'ID'
    ws['B3'].font = Font(color='0C4D75',bold=True)
    ws['B3'].fill = PatternFill(start_color="CCE9FC", end_color="CCE9FC", fill_type="solid")
    ws['C3'] = 'CEDULA'
    ws['C3'].font = Font(color='0C4D75', bold=True)
    ws['C3'].fill = PatternFill(start_color="CCE9FC", end_color="CCE9FC", fill_type="solid")
    ws['D3'] = 'APELLIDOS'
    ws['D3'].font = Font(color='0C4D75', bold=True)
    ws['D3'].fill = PatternFill(start_color="CCE9FC", end_color="CCE9FC", fill_type="solid")
    ws['E3'] = 'NOMBRES'
    ws['E3'].font = Font(color='0C4D75', bold=True)
    ws['E3'].fill = PatternFill(start_color="CCE9FC", end_color="CCE9FC", fill_type="solid")
    ws['F3'] = 'CORREO ELECTRONICO'
    ws['F3'].font = Font(color='0C4D75', bold=True)
    ws['F3'].fill = PatternFill(start_color="CCE9FC", end_color="CCE9FC", fill_type="solid")
    cell.font = cell.font.copy(bold=True)
    cont = 4
    for paciente in pacientes:
        ws.cell(row=cont, column=2).value = paciente.id
        ws.cell(row=cont, column=3).value = paciente.cedula
        ws.cell(row=cont, column=4).value = paciente.apellidos
        ws.cell(row=cont, column=5).value = paciente.nombres
        ws.cell(row=cont, column=6).value = paciente.correo
        cont = cont + 1
    nombre_archivo = "ReportePacientesMdClinic.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response



